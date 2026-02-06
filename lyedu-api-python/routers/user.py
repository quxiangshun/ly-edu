# -*- coding: utf-8 -*-
"""用户路由，与 Java UserController 对应"""
import io
from datetime import datetime as dt
from typing import Optional

from fastapi import APIRouter, File, Header, UploadFile
from openpyxl import load_workbook

from common.result import ResultCode, error, error_result, success
from models.schemas import UserRequest, ResetPasswordRequest
from services import user_service
from util.jwt_util import parse_authorization

router = APIRouter(prefix="/user", tags=["user"])


@router.get("/info")
def get_current_user(authorization: Optional[str] = Header(None, alias="Authorization")):
    """根据 token 返回当前用户信息，与前端 getUserInfo() 对接"""
    user_id = parse_authorization(authorization)
    if user_id is None:
        return error_result(ResultCode.UNAUTHORIZED)
    user = user_service.get_by_id(user_id)
    if not user:
        return error_result(ResultCode.USER_NOT_FOUND)
    return success({
        "id": user.get("id"),
        "username": user.get("username"),
        "realName": user.get("real_name"),
        "role": user.get("role") or "student",
    })


@router.get("/page")
def page(
    page: int = 1,
    size: int = 10,
    keyword: Optional[str] = None,
    departmentId: Optional[int] = None,
    role: Optional[str] = None,
    status: Optional[int] = None,
):
    return success(
        user_service.page(
            page_num=page,
            size=size,
            keyword=keyword,
            department_id=departmentId,
            role=role,
            status=status,
        )
    )


@router.get("/{id}")
def get_by_id(id: int):
    user = user_service.get_by_id(id)
    if not user:
        return error(404, "用户不存在")
    user.pop("password", None)
    return success(user)


@router.post("")
def create(body: UserRequest):
    if not body.username or not body.username.strip():
        return error(400, "用户名不能为空")
    existing = user_service.find_by_username(body.username.strip())
    if existing:
        return error(400, "用户名已存在")
    entry_d = None
    if body.entry_date and body.entry_date.strip():
        try:
            entry_d = dt.strptime(body.entry_date.strip()[:10], "%Y-%m-%d").date()
        except Exception:
            pass
    user_service.save(
        username=body.username.strip(),
        password=body.password,
        real_name=body.real_name,
        email=body.email,
        mobile=body.mobile,
        avatar=body.avatar,
        union_id=body.union_id,
        department_id=body.department_id,
        entry_date=entry_d,
        role=body.role or "student",
        status=body.status if body.status is not None else 1,
    )
    return success()


@router.put("/{id}")
def update(id: int, body: UserRequest):
    user = user_service.get_by_id(id)
    if not user:
        return error(404, "用户不存在")
    if body.username is not None and body.username.strip() != user.get("username"):
        existing = user_service.find_by_username(body.username.strip())
        if existing:
            return error(400, "用户名已存在")
    entry_d = None
    if body.entry_date is not None and body.entry_date and str(body.entry_date).strip():
        try:
            entry_d = dt.strptime(str(body.entry_date).strip()[:10], "%Y-%m-%d").date()
        except Exception:
            pass
    user_service.update(
        id,
        username=body.username.strip() if body.username else None,
        real_name=body.real_name,
        email=body.email,
        mobile=body.mobile,
        avatar=body.avatar,
        union_id=body.union_id,
        department_id=body.department_id,
        entry_date=entry_d,
        role=body.role,
        status=body.status,
    )
    return success()


@router.delete("/{id}")
def delete(id: int):
    user = user_service.get_by_id(id)
    if not user:
        return error(404, "用户不存在")
    user_service.delete(id)
    return success()


@router.post("/{id}/reset-password")
def reset_password(id: int, body: ResetPasswordRequest):
    user = user_service.get_by_id(id)
    if not user:
        return error(404, "用户不存在")
    user_service.update_password(id, body.password)
    return success()


# 员工导入模板表头（与前端下载模板一致）
ROLE_MAP = {"admin": "admin", "管理员": "admin", "teacher": "teacher", "教师": "teacher", "student": "student", "学员": "student"}


def _val_str(v) -> str:
    """从 Excel 单元格值转为字符串，支持数字、日期等"""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


@router.post("/import")
def import_users(file: UploadFile = File(...)):
    """上传 Excel 文件（.xlsx）批量导入员工。表头：用户名, 密码, 真实姓名, 邮箱, 手机号, 部门ID, 角色, 状态, 入职日期"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return error(400, "请上传 Excel 文件（.xlsx）")
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return error(400, f"文件读取失败: {str(e)}")
    try:
        ws = wb.active
        if not ws:
            return error(400, "文件为空")
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return error(400, "文件为空")
    # 首行为表头，数据从第 2 行起
    data_rows = rows[1:]
    success_count = 0
    messages = []
    for i, row in enumerate(data_rows):
        row_list = list(row) if row else []
        if not row_list or all(_val_str(v) == "" for v in row_list):
            continue
        while len(row_list) < 9:
            row_list.append("")
        row_num = i + 2
        username = _val_str(row_list[0])
        password = _val_str(row_list[1])
        real_name = _val_str(row_list[2]) or None
        email = _val_str(row_list[3]) or None
        mobile = _val_str(row_list[4]) or None
        dept_id = _val_str(row_list[5])
        role_str = (_val_str(row_list[6]) or "student").lower()
        status_str = _val_str(row_list[7])
        entry_date_str = _val_str(row_list[8])[:10]
        if not username:
            messages.append(f"第{row_num}行：用户名为空，已跳过")
            continue
        if user_service.find_by_username(username):
            messages.append(f"第{row_num}行：用户名「{username}」已存在，已跳过")
            continue
        department_id = None
        if dept_id and str(dept_id).replace(".0", "").isdigit():
            department_id = int(float(dept_id))
        role = ROLE_MAP.get(role_str) or ROLE_MAP.get(role_str.title()) or "student"
        status = 1
        if status_str in ("0", "禁用", "停用"):
            status = 0
        entry_d = None
        if entry_date_str:
            try:
                entry_d = dt.strptime(entry_date_str, "%Y-%m-%d").date()
            except Exception:
                pass
        try:
            user_service.save(
                username=username,
                password=password if password else None,
                real_name=real_name or None,
                email=email or None,
                mobile=mobile or None,
                department_id=department_id,
                entry_date=entry_d,
                role=role,
                status=status,
            )
            success_count += 1
        except Exception as e:
            messages.append(f"第{row_num}行：{username} 导入失败 - {str(e)}")
    return success({"successCount": success_count, "failCount": len(messages), "messages": messages})
