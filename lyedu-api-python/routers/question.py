# -*- coding: utf-8 -*-
"""试题路由，与 Java QuestionController 对应"""
import csv
import io
import json
from typing import Any, List, Optional, Tuple

from fastapi import APIRouter, File, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel

from common.result import error, error_result, success
from services.exam import question_service

router = APIRouter(prefix="/question", tags=["question"])

# 题型中文到英文映射（与前端一致）
TYPE_MAP = {
    "单选": "single",
    "多选": "multi",
    "判断": "judge",
    "填空": "fill",
    "简答": "short",
}


def _val_str(v) -> str:
    """从 Excel/单元格值转为字符串"""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _process_import_row(
    row_num: int,
    type_str: str,
    title: str,
    options_str: str,
    answer: Optional[str],
    score_str: str,
    analysis: Optional[str],
    sort_str: str,
) -> Tuple[bool, Optional[str]]:
    """处理单行导入数据，返回 (是否成功, 失败时的消息)。"""
    title = title.strip()
    if not title:
        return False, f"第{row_num}行：题干为空，已跳过"
    type_ = TYPE_MAP.get(type_str.strip())
    if not type_:
        return False, f"第{row_num}行：题型「{type_str}」无效，应为 单选/多选/判断/填空/简答，已跳过"
    score = 10
    if score_str and str(score_str).replace(".0", "").isdigit():
        score = max(1, min(100, int(float(score_str))))
    sort_val = 0
    if sort_str and str(sort_str).replace(".0", "").isdigit():
        sort_val = int(float(sort_str))
    options = (options_str.strip() or None) if options_str else None
    try:
        question_service.save(
            type_=type_,
            title=title,
            options=options,
            answer=answer,
            score=score,
            analysis=analysis,
            sort=sort_val,
        )
        return True, None
    except Exception as e:
        return False, f"第{row_num}行：{title[:20]}... 导入失败 - {str(e)}"


class QuestionRequest(BaseModel):
    type: str = ""
    title: str = ""
    options: Optional[str] = None
    answer: Optional[str] = None
    score: Optional[int] = 10
    analysis: Optional[str] = None
    sort: Optional[int] = 0


@router.get("/page")
def page(page: int = 1, size: int = 20, keyword: Optional[str] = None, type: Optional[str] = None):
    return success(question_service.page(page_num=page, size=size, keyword=keyword, type_=type))


@router.get("/{question_id}")
def get_by_id(question_id: int):
    q = question_service.get_by_id(question_id)
    if not q:
        return error_result((404, "资源不存在"))
    return success(q)


@router.post("")
def create(body: QuestionRequest):
    type_ = (body.type or "").strip()
    title = (body.title or "").strip()
    if not type_ or not title:
        return error(400, "题型和标题不能为空")
    qid = question_service.save(
        type_=type_,
        title=title,
        options=body.options,
        answer=body.answer,
        score=body.score or 10,
        analysis=body.analysis,
        sort=body.sort or 0,
    )
    return success(qid)


@router.put("/{question_id}")
def update(question_id: int, body: QuestionRequest):
    existing = question_service.get_by_id(question_id)
    if not existing:
        return error_result((404, "资源不存在"))
    type_ = (body.type or "").strip()
    title = (body.title or "").strip()
    if not type_ or not title:
        return error(400, "题型和标题不能为空")
    ok = question_service.update(
        question_id=question_id,
        type_=type_,
        title=title,
        options=body.options,
        answer=body.answer,
        score=body.score or 10,
        analysis=body.analysis,
        sort=body.sort or 0,
    )
    if not ok:
        return error(500, "更新失败")
    return success(None)


@router.delete("/{question_id}")
def delete(question_id: int):
    existing = question_service.get_by_id(question_id)
    if not existing:
        return error_result((404, "资源不存在"))
    question_service.delete(question_id)
    return success(None)


def _rows_from_xlsx(content: bytes) -> List[Tuple[int, List[Any]]]:
    """从 xlsx 内容解析出 (行号, 7 列列表)。首行为表头，从第 2 行起为数据。"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if not ws:
            return []
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    result = []
    for i, row in enumerate(rows[1:], start=2):
        row_list = list(row) if row else []
        if not row_list or all(_val_str(v) == "" for v in row_list):
            continue
        while len(row_list) < 7:
            row_list.append("")
        row_list = [_val_str(v) for v in row_list[:7]]
        result.append((i, row_list))
    return result


def _rows_from_csv(content: bytes) -> List[Tuple[int, List[str]]]:
    """从 CSV 内容解析出 (行号, 7 列列表)。首行为表头，UTF-8 支持 BOM。若某行列数多于 7（选项列含逗号被拆开），将中间列合并回选项列，保证参考答案/分值等对齐。"""
    try:
        text = content.decode("utf-8-sig")
    except Exception:
        text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    result = []
    for i, row in enumerate(rows[1:], start=2):
        row = [(s or "").strip() for s in row]
        if not row or all(s == "" for s in row):
            continue
        if len(row) > 7:
            merged = (
                [row[0], row[1], ",".join(row[2:-4])]
                + row[-4:]
            )
            row = merged
        while len(row) < 7:
            row.append("")
        row = row[:7]
        result.append((i, row))
    return result


def _rows_from_json(content: bytes) -> List[Tuple[int, List[Any]]]:
    """从 JSON 内容解析出 (行号, 7 列列表)。支持中文键或英文字段名。"""
    try:
        data = json.loads(content.decode("utf-8"))
    except Exception as e:
        raise ValueError(f"JSON 解析失败: {e}") from e
    if not isinstance(data, list):
        raise ValueError("JSON 根节点须为数组")
    # 键名映射：中文或英文 -> 列索引 0..6
    KEY_MAP = (
        ("题型", "type"),
        ("题干", "title"),
        ("选项(JSON)", "options"),
        ("参考答案", "answer"),
        ("分值", "score"),
        ("解析", "analysis"),
        ("排序", "sort"),
    )
    result = []
    for idx, item in enumerate(data):
        if not isinstance(item, dict):
            result.append((idx + 1, ["", "", "", "", "10", "", "0"]))
            continue
        row = []
        for cn, en in KEY_MAP:
            v = item.get(cn, item.get(en, ""))
            if v is None:
                v = ""
            if isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)
            row.append(str(v).strip())
        while len(row) < 7:
            row.append("")
        row = row[:7]
        result.append((idx + 1, row))
    return result


def _run_import_on_rows(rows_with_num: List[Tuple[int, List[Any]]]) -> dict:
    """对解析好的 (行号, 行数据) 列表执行导入，返回 successCount/failCount/messages。"""
    success_count = 0
    messages: List[str] = []
    for row_num, row_list in rows_with_num:
        type_str = row_list[0] if len(row_list) > 0 else ""
        title = row_list[1] if len(row_list) > 1 else ""
        options_str = row_list[2] if len(row_list) > 2 else ""
        answer = (row_list[3] or "").strip() or None if len(row_list) > 3 else None
        score_str = row_list[4] if len(row_list) > 4 else ""
        analysis = (row_list[5] or "").strip() or None if len(row_list) > 5 else None
        sort_str = row_list[6] if len(row_list) > 6 else ""
        ok, msg = _process_import_row(row_num, type_str, title, options_str, answer, score_str, analysis, sort_str)
        if ok:
            success_count += 1
        elif msg:
            messages.append(msg)
    return {"successCount": success_count, "failCount": len(messages), "messages": messages}


@router.post("/import")
def import_questions(file: UploadFile = File(...)):
    """上传 Excel(.xlsx) / CSV(.csv) / JSON(.json) 批量导入试题。表头/字段：题型, 题干, 选项(JSON), 参考答案, 分值, 解析, 排序"""
    fn = (file.filename or "").lower()
    if not fn.endswith(".xlsx") and not fn.endswith(".csv") and not fn.endswith(".json"):
        return error(400, "请上传 .xlsx、.csv 或 .json 文件")
    try:
        content = file.file.read()
    except Exception as e:
        return error(400, f"文件读取失败: {str(e)}")
    try:
        if fn.endswith(".xlsx"):
            rows_with_num = _rows_from_xlsx(content)
        elif fn.endswith(".csv"):
            rows_with_num = _rows_from_csv(content)
        else:
            try:
                rows_with_num = _rows_from_json(content)
            except ValueError as e:
                return error(400, str(e))
    except Exception as e:
        return error(400, f"文件解析失败: {str(e)}")
    return success(_run_import_on_rows(rows_with_num))


@router.post("/import/json")
def import_questions_json(body: List[dict]):
    """通过请求体直接提交 JSON 数组批量导入试题。每项为对象，支持中文键或英文字段名。"""
    if not isinstance(body, list):
        return error(400, "请求体须为 JSON 数组")
    try:
        rows_with_num = _rows_from_json(json.dumps(body, ensure_ascii=False).encode("utf-8"))
    except ValueError as e:
        return error(400, str(e))
    result = _run_import_on_rows(rows_with_num)
    return success(result)
